#!/Users/melocal/anaconda3/envs/c4g-wcag/bin python
import os
import sys
import time

import openpyxl
import pyautogui
import pynput
import selenium
from openpyxl import load_workbook
from pynput import keyboard
from pynput.keyboard import Controller, Key, KeyCode, Listener
from selenium import webdriver

# should print version 3.5 for pynput dependency
print("Python --version ==", sys.version)
print("openpyxl.__version__ ==", openpyxl.__version__)
print("selenium version ==", selenium.__version__)
# print("pynput version ==", pynput.__version__) # AttributeError: module 'pynput' has no attribute '__version__'

driver = webdriver.Safari(executable_path='/usr/bin/safaridriver')

os.chdir('/Users/melocal/MyDox/RedCross/Code4Good/WCAG2.0')
wb = openpyxl.load_workbook('RCB Top Sites B-2.xlsx')
sheets = wb.sheetnames
sheet0 = wb[sheets[0]]
print(sheet0.max_row)


def openAuditor(keyboard):
    print("inside openAuditor(keyboard)")
    keyboard.press(Key.alt)
    keyboard.press(Key.cmd)
    keyboard.press('i')
    keyboard.release('i')
    keyboard.release(Key.alt)
    keyboard.release(Key.cmd)


def captureKeyCombo(keyboard):
    ##################################################
    # https://github.com/moses-palmer/pynput/issues/20
    # The key combination to check
    COMBINATION = {keyboard.Key.cmd, keyboard.KeyCode(char='s')}

    # The currently active modifiers
    current = set()

    def on_press(key):
        try:
            if key in COMBINATION:
                current.add(key)
                if all(k in current for k in COMBINATION):
                    print(COMBINATION)
            if key == keyboard.Key.esc:
                listener.stop()
        except AttributeError as ex:
            print(ex)

    def on_release(key):
        try:
            current.remove(key)
            return True
        except KeyError:
            # pass
            return False

    with keyboard.Listener(on_press=on_press, on_release=on_release) as listener:
        listener.join()

    return COMBINATION


# for i in range(2, sheet0.max_row):
for i in range(49, 50):
    url = 'https://' + sheet0['A' + str(i)].value
    driver.get(url)
    print(i, driver.title)
    time.sleep(0.1)

    # pause the script to click around with mouse
    input()

    key_combo = captureKeyCombo(keyboard)
    print("key_combo ==", key_combo)
